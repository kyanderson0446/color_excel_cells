import os
from glob import glob
import openpyxl
from openpyxl.styles import PatternFill
path = r"C:\Users\kyle.anderson\Documents\Dan\Dan_pacs5"

for file in os.listdir(path):
    if ".xlsx" in file:
        full_path = os.path.join(path, file)
        wb = openpyxl.load_workbook(full_path)
        ws = wb['Sheet1'] # Name of the working sheet

        fill_cell1 = PatternFill(patternType='solid', fgColor='FCBA03')
        fill_cell2 = PatternFill(patternType='solid', fgColor='FCBA03')
        ws['E1'].fill = fill_cell1
        ws['K1'].fill = fill_cell2
        wb.save(fr"C:\Users\kyle.anderson\Documents\Dan\Dan_pacs5\{os.path.basename(file)}")

